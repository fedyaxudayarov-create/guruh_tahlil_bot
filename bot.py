import os, json, logging, re
from datetime import datetime, timezone, timedelta, date
from collections import defaultdict
from pathlib import Path

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, MessageHandler, CommandHandler,
    CallbackQueryHandler, filters, ContextTypes
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
#  SOZLAMALAR
# ═══════════════════════════════════════════════════════════════════
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
OWNER_ID  = int(os.getenv("OWNER_ID", "0"))
TZ        = timezone(timedelta(hours=5))          # Toshkent UTC+5

DATA_DIR  = Path(os.getenv("DATA_DIR", "/data"))
DATA_FILE = DATA_DIR / "logs.json"
MBRS_FILE = DATA_DIR / "members.json"

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════
#  STORAGE
# ═══════════════════════════════════════════════════════════════════
logs:    dict[str, list[dict]] = defaultdict(list)
members: dict[str, dict]       = defaultdict(dict)

def save_all():
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(dict(logs), f, ensure_ascii=False, default=str)
        with open(MBRS_FILE, "w", encoding="utf-8") as f:
            json.dump(dict(members), f, ensure_ascii=False, default=str)
    except Exception as e:
        log.error(f"save_all xatosi: {e}")

def load_all():
    global logs, members
    for path, target, name in [(DATA_FILE, "logs", "logs"), (MBRS_FILE, "members", "members")]:
        if path.exists():
            try:
                with open(path, encoding="utf-8") as f:
                    data = json.load(f)
                if name == "logs":
                    logs = defaultdict(list, data)
                else:
                    members = defaultdict(dict, data)
                log.info(f"✅ {name} yuklandi")
            except Exception as e:
                log.error(f"load {name} xatosi: {e}")

def reg_member(gid: str, uid: int, ism: str, username: str):
    members[gid][str(uid)] = {"ism": ism, "username": username}

# ═══════════════════════════════════════════════════════════════════
#  XABAR LINKI
# ═══════════════════════════════════════════════════════════════════
def msg_link(chat_id_str: str, msg_id: int) -> str:
    """Telegram xabar linkini qaytaradi"""
    if not msg_id:
        return ""
    try:
        cid = int(chat_id_str)
        abs_cid = str(abs(cid))
        # Supergroup: -1001234567890 → abs = 1001234567890 → [3:] = 1234567890
        if abs_cid.startswith("100") and len(abs_cid) >= 12:
            pure = abs_cid[3:]
        else:
            pure = abs_cid
        return f"https://t.me/c/{pure}/{msg_id}"
    except Exception:
        return ""

# ═══════════════════════════════════════════════════════════════════
#  TAHLIL YORDAMCHILARI
# ═══════════════════════════════════════════════════════════════════
def skript(t: str) -> str:
    if not t: return "-"
    k = len(re.findall(r"[а-яёА-ЯЁ]", t))
    l = len(re.findall(r"[a-zA-ZʻʼGgQqHh]", t))
    if k > 0 and l > 0: return "Aralash"
    if k > l: return "Kirill"
    if l > k: return "Lotin"
    return "-"

SALOM_KW = [
    "ассалому алайкум","салом","яхшимисиз",
    "assalomu alaykum","assalom","salom","yaxshimisiz",
]
TUSHUNDI_KW = [
    "тушунарли","tushunarli","тушундим","tushundim",
    "бажардим","bajardim","бажарилди","bajarildi",
    "хоп","hop","майли","mayli","ок","ok","👍","✅",
    "тайёр","tayyor","готово","тугади","tugadi","қабул","qabul",
]
XATO_KW = [
    "хато","xato","нотўғри","noto'g'ri","тушунмадим","tushunmadim",
    "билмадим","bilmadim","узр","uzr","кечирасиз","kechirasiz","❌",
]
SAVOL_KW = ["?","савол","savol","нима","nima","қандай","qanday","нега","nega"]

def has_salom(t): return any(k in t.lower() for k in SALOM_KW)
def holat(t):
    tl = t.lower()
    if any(k in tl for k in TUSHUNDI_KW): return "Tushundi"
    if any(k in tl for k in XATO_KW):    return "Xato"
    if any(k in tl for k in SAVOL_KW):   return "Savol"
    return "-"

def tahlil_msg(message) -> dict:
    user    = message.from_user
    chat    = message.chat
    caption = (message.caption or "").strip()
    now     = datetime.now(TZ)
    gid     = str(chat.id)
    mid     = message.message_id

    entry = {
        "ts"      : now.isoformat(),
        "sana"    : now.strftime("%Y-%m-%d"),
        "soat"    : now.strftime("%H:%M"),
        "guruh"   : chat.title or "Unknown",
        "guruh_id": gid,
        "ism"     : (f"{user.first_name or ''} {user.last_name or ''}".strip() if user else "Unknown"),
        "username": (f"@{user.username}" if user and user.username else str(user.id) if user else "-"),
        "user_id" : user.id if user else 0,
        "msg_id"  : mid,
        "link"    : msg_link(gid, mid),
        "tur"     : "", "izoh": "",
    }
    if   message.photo:      entry["tur"] = "rasm";    entry["izoh"] = caption
    elif message.video:      entry["tur"] = "video";   entry["izoh"] = caption
    elif message.video_note: entry["tur"] = "dumaloq"
    elif message.voice:      entry["tur"] = "ovoz";    entry["izoh"] = caption
    elif message.audio:      entry["tur"] = "audio";   entry["izoh"] = caption
    elif message.document:   entry["tur"] = "fayl";    entry["izoh"] = caption
    elif message.sticker:    entry["tur"] = "stiker"
    elif message.text:       entry["tur"] = "matn";    entry["izoh"] = message.text.strip()
    else:                    entry["tur"] = "boshqa"
    return entry

# ═══════════════════════════════════════════════════════════════════
#  EXCEL DIZAYN
# ═══════════════════════════════════════════════════════════════════
C = {
    "hdr"     : "1F4E79", "hdr_t"  : "FFFFFF",
    "hdr2"    : "C00000", "hdr2_t" : "FFFFFF",
    "juft"    : "EBF3FB", "toq"    : "FFFFFF",
    "green"   : "C6EFCE", "red_bg" : "FFE0E0",
    "yellow"  : "FFEB9C", "blue"   : "BDD7EE",
    "salom_bg": "E2EFDA", "link"   : "0563C1",
    "sum_bg"  : "2E4057", "sum_t"  : "FFFFFF",
}
chegara = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def pf(hex_): return PatternFill("solid", fgColor=hex_)

def sh(ws, row, col, val="", fon=None, bold=False, size=10,
       color="000000", wrap=True, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    if fon:  c.fill = pf(fon)
    c.font      = Font(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = chegara
    return c

# ─── Ustunlar ────────────────────────────────────────────────────
COLS = [
    ("№",              4),
    ("Ism",           24),
    ("Username",      16),
    ("Guruh",         20),
    ("📅 Birinchi\nxabar", 16),
    ("📅 Oxirgi\nxabar",   16),
    ("📷\nRasm",       6),
    ("🎥\nVideo",      6),
    ("⭕\nDumaloq",    7),
    ("🎤\nOvoz",       6),
    ("💬\nMatn",       6),
    ("📄\nFayl",       6),
    ("📦\nJami",       7),
    ("👋\nSalom",     10),
    ("✅\nTushundi",   9),
    ("❌\nXato",       7),
    ("❓\nSavol",      7),
    ("📝\nYozuv",      9),
    ("Xabarlar (matn • link)",  60),
]
SARLAVHA = [c[0] for c in COLS]
KENGLIK  = [c[1] for c in COLS]

# ─── Yuborganlar sheet ───────────────────────────────────────────
def sheet_yuborganlar(wb, nom: str, glog: list, sana_filter: str | None = None) -> set:
    title = f"✅ {nom}"[:30]
    ws = wb.create_sheet(title=title)

    # Agar sana filter bo'lsa — sarlavhaga yozamiz
    if sana_filter:
        ws.cell(row=1, column=1, value=f"📅 Sana: {sana_filter}").font = Font(
            bold=True, size=10, color="1F4E79"
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        ws.row_dimensions[1].height = 18
        hdr_row = 2
    else:
        hdr_row = 1

    # Sarlavha qatori
    for col, s in enumerate(SARLAVHA, 1):
        sh(ws, hdr_row, col, s, fon=C["hdr"], bold=True, color=C["hdr_t"])
    ws.row_dimensions[hdr_row].height = 34

    # Faqat tanlangan sanani filter qilish
    if sana_filter:
        glog = [l for l in glog if l.get("sana") == sana_filter]

    # Foydalanuvchi ma'lumotlarini jamlash
    a: dict[int, dict] = {}
    for l in sorted(glog, key=lambda x: x.get("ts", "")):
        uid = l["user_id"]
        if uid not in a:
            a[uid] = {
                "ism": l["ism"], "username": l["username"], "guruh": l["guruh"],
                "birinchi": l["soat"],
                "oxirgi"  : l["soat"],
                "birinchi_full": l["sana"] + " " + l["soat"],
                "oxirgi_full"  : l["sana"] + " " + l["soat"],
                "rasm":0,"video":0,"dumaloq":0,"ovoz":0,
                "matn":0,"fayl":0,
                "salom":0,"tushundi":0,"xato":0,"savol":0,
                "lotin":0,"kirill":0,"aralash":0,
                "xabarlar": [],   # (matn, link) juftliklari
            }
        d = a[uid]
        d["oxirgi_full"] = l["sana"] + " " + l["soat"]

        t = l["tur"]
        if   t == "rasm":            d["rasm"]    += 1
        elif t == "video":           d["video"]   += 1
        elif t == "dumaloq":         d["dumaloq"] += 1
        elif t == "ovoz":            d["ovoz"]    += 1
        elif t == "matn":            d["matn"]    += 1
        elif t in ("fayl","audio"):  d["fayl"]    += 1

        izoh = l.get("izoh", "").strip()
        link = l.get("link", "")
        if izoh or link:
            if izoh and has_salom(izoh): d["salom"] += 1
            h = holat(izoh) if izoh else "-"
            if   h == "Tushundi": d["tushundi"] += 1
            elif h == "Xato":     d["xato"]     += 1
            elif h == "Savol":    d["savol"]    += 1
            if izoh:
                s_ = skript(izoh)
                if   "Lotin"   in s_: d["lotin"]   += 1
                elif "Kirill"  in s_: d["kirill"]  += 1
                elif "Aralash" in s_: d["aralash"] += 1
            qisqa = (izoh[:100] + "…") if len(izoh) > 100 else izoh
            d["xabarlar"].append((qisqa or f"[{t}]", link, l["soat"]))

    # Ma'lumot qatorlari
    data_start = hdr_row + 1
    for i, (uid, d) in enumerate(
        sorted(a.items(), key=lambda x: -(
            x[1]["rasm"]+x[1]["video"]+x[1]["dumaloq"]+
            x[1]["ovoz"]+x[1]["matn"]+x[1]["fayl"]
        )), 1
    ):
        row  = data_start + i - 1
        fon  = C["juft"] if i % 2 == 0 else C["toq"]
        jami = d["rasm"]+d["video"]+d["dumaloq"]+d["ovoz"]+d["matn"]+d["fayl"]

        yozuv = max(
            [("Lotin",d["lotin"]),("Kirill",d["kirill"]),("Aralash",d["aralash"])],
            key=lambda x: x[1]
        )[0] if (d["lotin"] or d["kirill"] or d["aralash"]) else "-"

        # Xabarlar ustuni: "soat • matn → link"
        xabar_lines = []
        for matn, link, soat in d["xabarlar"]:
            if link:
                xabar_lines.append(f"{soat} • {matn}\n  🔗 {link}")
            else:
                xabar_lines.append(f"{soat} • {matn}")
        xabar_cell = "\n\n".join(xabar_lines) if xabar_lines else "-"

        qiymatlar = [
            i,
            d["ism"],
            d["username"],
            d["guruh"],
            d["birinchi_full"],
            d["oxirgi_full"],
            d["rasm"]    or "",
            d["video"]   or "",
            d["dumaloq"] or "",
            d["ovoz"]    or "",
            d["matn"]    or "",
            d["fayl"]    or "",
            jami,
            d["salom"]   or "",
            d["tushundi"]or "",
            d["xato"]    or "",
            d["savol"]   or "",
            yozuv,
            xabar_cell,
        ]
        for col, q in enumerate(qiymatlar, 1):
            c = sh(ws, row, col, q, fon=fon)
            if col == 19:   # Xabarlar ustuni
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.font = Font(size=9)

        # Rangli belgilash
        if d["salom"]    > 0: ws.cell(row=row, column=14).fill = pf(C["salom_bg"])
        if d["tushundi"] > 0: ws.cell(row=row, column=15).fill = pf(C["green"])
        if d["xato"]     > 0: ws.cell(row=row, column=16).fill = pf(C["red_bg"])
        if d["savol"]    > 0: ws.cell(row=row, column=17).fill = pf(C["yellow"])
        for mc in [7,8,9,10,11,12]:
            if ws.cell(row=row, column=mc).value:
                ws.cell(row=row, column=mc).fill = pf(C["blue"])

        # Qator balandligi
        n_lines = max(1, len(xabar_lines))
        ws.row_dimensions[row].height = min(max(30, n_lines * 28), 250)

    # Ustun kengligi
    for col, k in enumerate(KENGLIK, 1):
        ws.column_dimensions[get_column_letter(col)].width = k
    ws.freeze_panes = f"A{data_start}"

    # ─── Pastki JAMI qatori (bitta, sodda) ──────────────────────
    jami_row = data_start + len(a)
    ws.row_dimensions[jami_row].height = 22
    jami_vals = {
        1:  f"Jami: {len(a)} kishi",
        7:  sum(d["rasm"]    for d in a.values()),
        8:  sum(d["video"]   for d in a.values()),
        9:  sum(d["dumaloq"] for d in a.values()),
        10: sum(d["ovoz"]    for d in a.values()),
        11: sum(d["matn"]    for d in a.values()),
        12: sum(d["fayl"]    for d in a.values()),
        13: sum(d["rasm"]+d["video"]+d["dumaloq"]+d["ovoz"]+d["matn"]+d["fayl"]
                for d in a.values()),
        14: sum(1 for d in a.values() if d["salom"] > 0),
        15: sum(1 for d in a.values() if d["tushundi"] > 0),
    }
    for col in range(1, len(COLS)+1):
        val = jami_vals.get(col, "")
        sh(ws, jami_row, col, val, fon=C["sum_bg"], bold=True,
           color=C["sum_t"], size=9)

    # Hisobot vaqti
    ws.cell(row=jami_row+1, column=1,
            value=f"Hisobot: {datetime.now(TZ).strftime('%Y-%m-%d %H:%M')}   |   "
                  f"Guruh: {nom}   |   Jami xabarlar: {len(glog)}"
    ).font = Font(italic=True, size=8, color="888888")

    return set(a.keys())   # xabar yuborgan uid'lar


# ─── Bermaganlar sheet ───────────────────────────────────────────
def sheet_bermaganlar(wb, gid: str, nom: str, yuborgan_uids: set):
    ws = wb.create_sheet(title=f"❌ Bermaganlar"[:30])

    # Sarlavha
    for col, (s, _) in enumerate([("№",4),("Ism",30),("Username",20),("Holati",20)], 1):
        sh(ws, 1, col, s, fon=C["hdr2"], bold=True, color=C["hdr2_t"])
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    gmbrs      = members.get(gid, {})
    bermaganlar = {uid: info for uid, info in gmbrs.items()
                   if int(uid) not in yuborgan_uids}

    if not bermaganlar:
        c = ws.cell(row=2, column=1, value="✅ Barcha a'zolar xabar yuborgan!")
        c.font = Font(bold=True, size=11, color="00AA00")
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:D2")
    else:
        for i, (uid, info) in enumerate(bermaganlar.items(), 1):
            r = i + 1
            sh(ws, r, 1, i,                  fon=C["red_bg"])
            sh(ws, r, 2, info["ism"],        fon=C["red_bg"], align="left")
            sh(ws, r, 3, info["username"],   fon=C["red_bg"])
            sh(ws, r, 4, "❌ Xabar bermagan", fon=C["red_bg"])
            ws.row_dimensions[r].height = 18

    # Jami
    jr = len(bermaganlar) + 3
    c = ws.cell(row=jr, column=1,
                value=f"Jami: {len(gmbrs)} a'zo  |  "
                      f"Xabar bergan: {len(yuborgan_uids)}  |  "
                      f"Bermagan: {len(bermaganlar)}")
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = pf(C["hdr2"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=jr, start_column=1, end_row=jr, end_column=4)
    ws.row_dimensions[jr].height = 22

    ws.cell(row=jr+1, column=1,
            value="* Bot qo'shilgandan beri ko'rilgan a'zolar asosida"
    ).font = Font(italic=True, size=8, color="888888")
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════
#  EXCEL YUBORISH
# ═══════════════════════════════════════════════════════════════════
async def send_excel(ctx, chat_id: int, target_gids: list[str], sana: str | None = None):
    tahlil_data = {gid: logs[gid] for gid in target_gids if gid in logs and logs[gid]}
    if not tahlil_data:
        await ctx.bot.send_message(chat_id=chat_id, text="📭 Ma'lumot yo'q.")
        return

    await ctx.bot.send_message(chat_id=chat_id, text="⏳ Excel tayyorlanmoqda…")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for gid, glog in tahlil_data.items():
        nom = glog[-1]["guruh"]
        yuborgan = sheet_yuborganlar(wb, nom, glog, sana_filter=sana)
        sheet_bermaganlar(wb, gid, nom, yuborgan)

    fayl_nomi = f"tahlil_{datetime.now(TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    path      = f"/tmp/{fayl_nomi}"
    wb.save(path)

    guruh_nomlari = ", ".join(logs[gid][-1]["guruh"] for gid in tahlil_data if logs[gid])
    caption = (
        f"📊 *Excel tayyor!*\n"
        f"🏷 {guruh_nomlari}\n"
        + (f"📅 Sana: {sana}\n" if sana else "")
        + f"📋 Sheet 1: ✅ Xabar yuborganlar\n"
          f"📋 Sheet 2: ❌ Xabar bermaganlar"
    )
    with open(path, "rb") as f:
        await ctx.bot.send_document(
            chat_id=chat_id, document=f,
            filename=fayl_nomi, caption=caption, parse_mode="Markdown",
        )
    os.remove(path)


# ═══════════════════════════════════════════════════════════════════
#  INLINE KEYBOARD YORDAMCHILARI
# ═══════════════════════════════════════════════════════════════════
def kb_guruhlar() -> InlineKeyboardMarkup | None:
    if not any(logs.values()): return None
    rows = []
    for gid, glog in logs.items():
        if not glog: continue
        nom  = glog[-1]["guruh"]
        nxbr = len(glog)
        uids = len(set(l["user_id"] for l in glog))
        rows.append([InlineKeyboardButton(
            f"📁 {nom[:30]} ({nxbr} xabar, {uids} kishi)",
            callback_data=f"grp:{gid}"
        )])
    rows.append([InlineKeyboardButton("📊 BARCHA guruhlar", callback_data="grp:ALL")])
    return InlineKeyboardMarkup(rows) if rows else None

def kb_sana(gid: str) -> InlineKeyboardMarkup:
    today     = datetime.now(TZ).strftime("%Y-%m-%d")
    yesterday = (datetime.now(TZ) - timedelta(days=1)).strftime("%Y-%m-%d")
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"📅 Bugun ({today})",      callback_data=f"date:{gid}:{today}")],
        [InlineKeyboardButton(f"📅 Kecha ({yesterday})",  callback_data=f"date:{gid}:{yesterday}")],
        [InlineKeyboardButton("📅 Boshqa sana (YYYY-MM-DD)", callback_data=f"date:{gid}:custom")],
        [InlineKeyboardButton("📊 Barchasi (sana filtrsiz)", callback_data=f"date:{gid}:all")],
    ])


# ═══════════════════════════════════════════════════════════════════
#  HANDLERS
# ═══════════════════════════════════════════════════════════════════
async def guruh_xabar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = update.message or update.channel_post
    if not msg or not msg.from_user: return
    entry = tahlil_msg(msg)
    gid   = entry["guruh_id"]
    logs[gid].append(entry)
    reg_member(gid, entry["user_id"], entry["ism"], entry["username"])
    if len(logs[gid]) % 20 == 0:
        save_all()

async def yangi_azolar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg or not msg.new_chat_members: return
    gid = str(msg.chat.id)
    for u in msg.new_chat_members:
        if u.is_bot: continue
        ism = f"{u.first_name or ''} {u.last_name or ''}".strip()
        reg_member(gid, u.id, ism, f"@{u.username}" if u.username else str(u.id))
    save_all()

async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    await update.message.reply_text(
        "🤖 *Guruh Tahlil Boti v4*\n\n"
        "📋 *Buyruqlar:*\n"
        "/report — Guruh → Sana tanlash → Excel\n"
        "/stats — Guruh statistikasi\n"
        "/guruhlar — Guruhlar ro'yxati\n"
        "/save — Qo'lda saqlash\n"
        "/clear — Ma'lumotlarni tozalash",
        parse_mode="Markdown",
    )

async def report_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    kb = kb_guruhlar()
    if kb is None:
        await update.message.reply_text("📭 Hozircha ma'lumot yo'q.")
        return
    await update.message.reply_text(
        "📊 *1-qadam: Guruhni tanlang*",
        reply_markup=kb, parse_mode="Markdown",
    )

async def stats_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    if not any(logs.values()):
        await update.message.reply_text("📭 Ma'lumot yo'q."); return
    aktiv = [g for g in logs if logs[g]]
    if len(aktiv) == 1:
        await _show_stats(update.message, aktiv[0]); return
    rows = []
    for gid, glog in logs.items():
        if not glog: continue
        nom = glog[-1]["guruh"]
        rows.append([InlineKeyboardButton(f"📊 {nom[:35]}", callback_data=f"stats:{gid}")])
    await update.message.reply_text(
        "📊 Qaysi guruh statistikasi?",
        reply_markup=InlineKeyboardMarkup(rows),
    )

async def _show_stats(msg_or_query, gid: str):
    glog = logs.get(gid, [])
    if not glog:
        text = "📭 Bu guruhda ma'lumot yo'q."
    else:
        nom  = glog[-1]["guruh"]
        uids = len(set(l["user_id"] for l in glog))
        mbrs = len(members.get(gid, {}))
        text = (
            f"📊 *{nom}*\n\n"
            f"👥 Ko'rilgan a'zolar: *{mbrs}*\n"
            f"✅ Xabar yuborganlar: *{uids}*\n"
            f"❌ Xabar bermaganlar: *{max(0, mbrs-uids)}*\n"
            f"📨 Jami xabarlar: *{len(glog)}*\n\n"
            f"📷 Rasm: {sum(1 for l in glog if l['tur']=='rasm')}  "
            f"🎥 Video: {sum(1 for l in glog if l['tur']=='video')}  "
            f"⭕ Dumaloq: {sum(1 for l in glog if l['tur']=='dumaloq')}\n"
            f"🎤 Ovoz: {sum(1 for l in glog if l['tur']=='ovoz')}  "
            f"💬 Matn: {sum(1 for l in glog if l['tur']=='matn')}\n\n"
            f"📅 Oxirgi: {glog[-1]['sana']} {glog[-1]['soat']}"
        )
    if hasattr(msg_or_query, "reply_text"):
        await msg_or_query.reply_text(text, parse_mode="Markdown")
    else:
        await msg_or_query.edit_message_text(text, parse_mode="Markdown")

async def guruhlar_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    if not any(logs.values()):
        await update.message.reply_text("📭 Hech qanday guruh yo'q."); return
    javob = "👥 *Guruhlar:*\n\n"
    for gid, glog in logs.items():
        if not glog: continue
        nom  = glog[-1]["guruh"]
        uids = len(set(l["user_id"] for l in glog))
        mbrs = len(members.get(gid, {}))
        javob += (f"🏷 *{nom}*\n"
                  f"  👥 {mbrs} a'zo  ✅ {uids} faol  ❌ {max(0,mbrs-uids)} bermagan\n"
                  f"  ID: `{gid}`\n\n")
    await update.message.reply_text(javob, parse_mode="Markdown")

async def save_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    save_all()
    await update.message.reply_text("💾 Saqlandi.")

async def clear_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Ha, tozala", callback_data="clear:yes"),
        InlineKeyboardButton("❌ Bekor",      callback_data="clear:no"),
    ]])
    await update.message.reply_text(
        "⚠️ *Barcha ma'lumotlar o'chadi!*", reply_markup=kb, parse_mode="Markdown"
    )

# ─── Matnli xabar → custom sana kiritish ────────────────────────
async def matn_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    pending = ctx.user_data.get("pending_date")
    if not pending: return

    matn = (update.message.text or "").strip()
    # Format tekshirish: YYYY-MM-DD
    try:
        datetime.strptime(matn, "%Y-%m-%d")
    except ValueError:
        await update.message.reply_text(
            "❌ Format noto'g'ri. Iltimos: `YYYY-MM-DD`\nMasalan: `2026-05-20`",
            parse_mode="Markdown",
        )
        return

    gid_val = pending["gid"]
    ctx.user_data.pop("pending_date")

    target = list(logs.keys()) if gid_val == "ALL" else [gid_val]
    await send_excel(ctx, OWNER_ID, target, sana=matn)

# ─── Callback ────────────────────────────────────────────────────
async def callback_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.from_user.id != OWNER_ID:
        await q.answer("❌ Ruxsat yo'q", show_alert=True); return

    data = q.data

    # 1-qadam: guruh tanlash
    if data.startswith("grp:"):
        gid_val = data[4:]
        await q.edit_message_text(
            "📅 *2-qadam: Sanani tanlang*",
            reply_markup=kb_sana(gid_val),
            parse_mode="Markdown",
        )

    # 2-qadam: sana tanlash
    elif data.startswith("date:"):
        _, gid_val, sana_val = data.split(":", 2)
        if sana_val == "custom":
            ctx.user_data["pending_date"] = {"gid": gid_val}
            await q.edit_message_text(
                "✏️ Sanani yozing (format: `YYYY-MM-DD`)\nMasalan: `2026-05-20`",
                parse_mode="Markdown",
            )
            return
        target = list(logs.keys()) if gid_val == "ALL" else [gid_val]
        sana   = None if sana_val == "all" else sana_val
        await q.edit_message_text("⏳ Excel tayyorlanmoqda…")
        await send_excel(ctx, OWNER_ID, target, sana=sana)

    # Stats
    elif data.startswith("stats:"):
        await _show_stats(q, data[6:])

    # Clear
    elif data == "clear:yes":
        logs.clear(); members.clear(); save_all()
        await q.edit_message_text("🗑 Tozalandi.")
    elif data == "clear:no":
        await q.edit_message_text("✅ Bekor qilindi.")


# ═══════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════
def main():
    if not BOT_TOKEN: raise ValueError("❌ BOT_TOKEN yo'q!")
    if not OWNER_ID:  raise ValueError("❌ OWNER_ID yo'q!")
    load_all()

    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start",    start))
    app.add_handler(CommandHandler("report",   report_cmd))
    app.add_handler(CommandHandler("stats",    stats_cmd))
    app.add_handler(CommandHandler("guruhlar", guruhlar_cmd))
    app.add_handler(CommandHandler("save",     save_cmd))
    app.add_handler(CommandHandler("clear",    clear_cmd))
    app.add_handler(CallbackQueryHandler(callback_handler))

    # Yangi a'zolar
    app.add_handler(MessageHandler(
        filters.ChatType.GROUPS & filters.StatusUpdate.NEW_CHAT_MEMBERS,
        yangi_azolar
    ))
    # Owner — shaxsiy chat (custom sana kiritish)
    app.add_handler(MessageHandler(
        filters.ChatType.PRIVATE & filters.TEXT & filters.User(OWNER_ID),
        matn_handler
    ))
    # Guruh xabarlari
    app.add_handler(MessageHandler(
        filters.ChatType.GROUPS & (
            filters.TEXT | filters.PHOTO | filters.VIDEO |
            filters.VIDEO_NOTE | filters.VOICE | filters.AUDIO |
            filters.Document.ALL | filters.Sticker.ALL
        ),
        guruh_xabar
    ))

    log.info("🤖 Bot v4 ishga tushdi")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
